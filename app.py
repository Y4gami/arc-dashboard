import sqlite3
import io
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename

app = Flask(__name__)
DATABASE = 'arc_dashboard.db'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS arcs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula TEXT UNIQUE NOT NULL,
            tipo_operacion TEXT NOT NULL DEFAULT '',
            sn TEXT NOT NULL DEFAULT '',
            modelo TEXT NOT NULL,
            fecha_arc TEXT NOT NULL,
            fecha_proximo_arc TEXT NOT NULL,
            tipo_arc TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'Sin Iniciar',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/arcs', methods=['GET', 'POST'])
def api_arcs():
    conn = get_db()
    if request.method == 'POST':
        data = request.json
        try:
            conn.execute('''
                INSERT INTO arcs (matricula, tipo_operacion, sn, modelo, fecha_arc, fecha_proximo_arc, tipo_arc, estado)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (data['matricula'], data['tipo_operacion'], data['sn'], data['modelo'],
                  data['fecha_arc'], data['fecha_proximo_arc'], data['tipo_arc'], data['estado']))
            conn.commit()
            return jsonify({'status': 'ok'}), 201
        except sqlite3.IntegrityError:
            return jsonify({'status': 'error', 'message': 'Matrícula ya existe'}), 400
        finally:
            conn.close()

    arcs = conn.execute('SELECT * FROM arcs ORDER BY fecha_proximo_arc ASC').fetchall()
    conn.close()
    return jsonify([dict(row) for row in arcs])

@app.route('/api/arcs/<int:arc_id>', methods=['GET', 'PUT', 'DELETE'])
def api_arc_edit(arc_id):
    conn = get_db()
    if request.method == 'GET':
        arc = conn.execute('SELECT * FROM arcs WHERE id=?', (arc_id,)).fetchone()
        conn.close()
        if arc:
            return jsonify(dict(arc))
        return jsonify({'status': 'error', 'message': 'No encontrado'}), 404

    if request.method == 'PUT':
        data = request.json
        conn.execute('''
            UPDATE arcs SET matricula=?, tipo_operacion=?, sn=?, modelo=?, fecha_arc=?, fecha_proximo_arc=?, tipo_arc=?, estado=?
            WHERE id=?
        ''', (data['matricula'], data['tipo_operacion'], data['sn'], data['modelo'],
              data['fecha_arc'], data['fecha_proximo_arc'], data['tipo_arc'], data['estado'], arc_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})

    if request.method == 'DELETE':
        conn.execute('DELETE FROM arcs WHERE id=?', (arc_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'ok'})

@app.route('/api/import', methods=['POST'])
def api_import():
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'status': 'error', 'message': 'Formato no válido. Usa .xlsx o .xls'}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        conn = get_db()
        imported = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                matricula = str(row[0]).strip() if row[0] else None
                tipo_operacion = str(row[1]).strip() if row[1] else None
                sn = str(row[2]).strip() if row[2] else None
                modelo = str(row[3]).strip() if row[3] else None
                fecha_arc = str(row[4]).strip() if row[4] else None
                fecha_proximo_arc = str(row[5]).strip() if row[5] else None
                tipo_arc = str(row[6]).strip() if row[6] else None
                estado = str(row[7]).strip() if len(row) > 7 and row[7] else 'Sin Iniciar'

                if not all([matricula, tipo_operacion, sn, modelo, fecha_arc, fecha_proximo_arc, tipo_arc]):
                    errors.append(f'Fila {row_idx}: datos incompletos')
                    continue

                # Normalize date formats
                for date_field in [fecha_arc, fecha_proximo_arc]:
                    if date_field and isinstance(date_field, str):
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
                            try:
                                dt = datetime.strptime(date_field, fmt)
                                if date_field == fecha_arc:
                                    fecha_arc = dt.strftime('%Y-%m-%d')
                                else:
                                    fecha_proximo_arc = dt.strftime('%Y-%m-%d')
                                break
                            except:
                                pass

                conn.execute('''
                    INSERT OR REPLACE INTO arcs (matricula, tipo_operacion, sn, modelo, fecha_arc, fecha_proximo_arc, tipo_arc, estado)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (matricula, tipo_operacion, sn, modelo, fecha_arc, fecha_proximo_arc, tipo_arc, estado))
                imported += 1
            except Exception as e:
                errors.append(f'Fila {row_idx}: {str(e)}')

        conn.commit()
        conn.close()
        wb.close()

        return jsonify({'status': 'ok', 'imported': imported, 'errors': errors})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/export', methods=['GET'])
def api_export():
    try:
        import openpyxl
        from io import BytesIO

        conn = get_db()
        arcs = conn.execute('SELECT * FROM arcs ORDER BY fecha_proximo_arc ASC').fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'ARCs'

        headers = ['MATRÍCULA', 'TIPO OPERACIÓN', 'SN', 'MODELO', 'Fecha ARC', 'Fecha PRÓXIMO ARC', 'TIPO ARC', 'ESTADO']
        ws.append(headers)

        for arc in arcs:
            ws.append([
                arc['matricula'],
                arc['tipo_operacion'],
                arc['sn'],
                arc['modelo'],
                arc['fecha_arc'],
                arc['fecha_proximo_arc'],
                arc['tipo_arc'],
                arc['estado']
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       as_attachment=True, download_name='ARCs_export.xlsx')

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=False)
